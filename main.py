from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

Url = "https://www.pcone.com.tw/"
Driver = webdriver.Edge()
Driver.get(Url)
time.sleep(5)
# 向下捲動頁面
for i in range(15):
    Driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
    time.sleep(2)
PageContent = Driver.page_source # 取得頁面內容
Html = BeautifulSoup(PageContent, 'html.parser')
Products = Html.find(class_ = "infinite-scroll-component__outerdiv").findChild()
ProductsInfo = Products.find_all(class_ = "product-info")
ProductsName = []
ProductsDiscount = []
ProductsPrice = []
ProductsStar = []
ProductsSold = []

# 建立excel檔
OutputExcel = "shoplist.xlsx"
try:
    Wb = openpyxl.load_workbook(OutputExcel)
    Sheet = Wb["商品清單"]
except FileNotFoundError:
    Wb = openpyxl.Workbook()
    Wb.create_sheet("商品清單", 0)
    Sheet = Wb["商品清單"]
    for index, value in [("A", "商品名稱"), ("B", "折扣"), ("C", "價錢"), ("D", "評價"), ("E", "搶購人數")]:
        Sheet[f"{index}1"].value = value
    Wb.save(OutputExcel)

# 把爬蟲資料寫入excel
for index, value in enumerate(ProductsInfo):
    ProductsName.append(value.find(class_ = "product-name"))
    ProductsDiscount.append(value.find(class_ = "label-wording"))
    ProductsPrice.append(value.find(class_ = "price"))
    ProductsStar.append(value.find(class_ = "review-avg"))
    ProductsSold.append(value.find(class_ = "sold-wording"))
    # 分別判斷是否為None並使用get_text()取值
    for index2, value2 in [
        ("A", ProductsName[index]),
        ("B", ProductsDiscount[index]),
        ("C", ProductsPrice[index]),
        ("D", ProductsStar[index]),
        ("E", ProductsSold[index])
    ]:
        if value2 != None:
            print(value2.get_text())
            Sheet[f"{index2}{index + 2}"].value = f"{value2.get_text()}"
        else:
            print("None")
            Sheet[f"{index2}{index + 2}"].value = "None"

Wb.save(OutputExcel)
Wb.close()